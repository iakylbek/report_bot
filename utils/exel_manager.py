from datetime import datetime

from openpyxl import load_workbook, Workbook

from config import FILE_PATH


class ExelManager():
    _instance = None

    workbook = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super(ExelManager, cls).__new__(cls)
        return cls._instance

    def __init__(self):
        if not hasattr(self, '_initialized'):
            try:
                self.workbook = load_workbook(FILE_PATH)
            except:
                self.workbook = Workbook()
                self.workbook.remove(self.workbook.active)

                self._create_sheet(
                    "Cотрудники", ["Telegram ID", "Имя Пользователя"])
                self._create_sheet(
                    "Товары", ["Артикул", "Наименование", "Цена", "Кто отвественный", "Комментарий"])
                self._create_sheet(
                    "Остатки", ["Ответственный", "Артикул",
                                "Товар", "Остаток", "Общая стоимость"]
                )
                self._create_sheet(
                    "Операции", ["Дата", "Товар", "Поступление or отгрузка", "Количество", "Комментарий"])

                self.save_data()

            self._initialized = True

    def _create_sheet(self, name: str, headers):
        self.workbook.create_sheet(title=f"{name}")
        sheet = self.workbook[name]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

    def save_data(self):
        self.workbook.save(FILE_PATH)

    def add_employee(self, telegram_id, name):
        sheet = self.workbook["Cотрудники"]
        next_row = sheet.max_row + 1

        sheet.cell(row=next_row, column=1, value=telegram_id)
        sheet.cell(row=next_row, column=2, value=name)

        self.save_data()

    def add_product(self, article, name, price, responsible, comment):
        sheet = self.workbook["Товары"]
        next_row = sheet.max_row + 1

        sheet.cell(row=next_row, column=1, value=article)
        sheet.cell(row=next_row, column=2, value=name)
        sheet.cell(row=next_row, column=3, value=price)
        sheet.cell(row=next_row, column=4, value=responsible)
        sheet.cell(row=next_row, column=5, value=comment)

        self.save_data()

    def add_operation(self, product_name, op_type, quantity, comment):
        sheet = self.workbook["Операции"]
        next_row = sheet.max_row + 1

        date_str = datetime.now().strftime("%d.%m.%Y")

        sheet.cell(row=next_row, column=1, value=date_str)
        sheet.cell(row=next_row, column=2, value=product_name)
        sheet.cell(row=next_row, column=3, value=op_type)
        sheet.cell(row=next_row, column=4, value=quantity)
        sheet.cell(row=next_row, column=5, value=comment)

        self._update_stock(product_name, op_type, quantity)
        self.save_data()

    def _update_stock(self, product_name, op_type, quantity):
        sheet_products = self.workbook["Товары"]
        sheet_stock = self.workbook["Остатки"]

        product_row = None
        for i in range(2, sheet_products.max_row + 1):
            if sheet_products.cell(i, 2).value == product_name:
                product_row = i
                break

        if not product_row:
            return  # Товар не найден

        article = sheet_products.cell(product_row, 1).value
        price = int(sheet_products.cell(product_row, 3).value) or 0
        responsible = sheet_products.cell(product_row, 4).value or "-"

        # Ищем товар в таблице "Остатки"
        stock_row = None
        for i in range(2, sheet_stock.max_row + 1):
            if sheet_stock.cell(i, 3).value == product_name:
                stock_row = i
                break

        # Определяем изменение остатка
        delta = quantity if op_type == "Поступление" else -quantity

        if stock_row:
            # Обновляем существующий товар в "Остатки"
            current_stock = sheet_stock.cell(stock_row, 4).value or 0
            new_stock = current_stock + delta

            # Обновляем строку остатка товара
            # Новый остаток
            sheet_stock.cell(stock_row, 4, new_stock)
            # Новая общая стоимость
            sheet_stock.cell(stock_row, 5, new_stock * price)
        else:
            # Добавляем новый товар в "Остатки" (если его еще нет)
            new_row = sheet_stock.max_row + 1
            initial_stock = quantity if op_type == "Поступление" else -quantity

            sheet_stock.cell(new_row, 1, responsible)
            sheet_stock.cell(new_row, 2, article)
            sheet_stock.cell(new_row, 3, product_name)
            sheet_stock.cell(new_row, 4, initial_stock)          # Остаток
            sheet_stock.cell(new_row, 5, initial_stock *
                             price)  # Общая стоимость

    def get_all_remains(self):
        sheet_stock = self.workbook["Остатки"]
        all_products = []

        for row in range(2, sheet_stock.max_row + 1):
            product_name = sheet_stock.cell(row, 3).value  # Название товара
            remaining_quantity = sheet_stock.cell(
                row, 4).value
            total_value = sheet_stock.cell(row, 5).value  # Общая стоимость

            all_products.append(
                f"Товар: {product_name}, Количество: {remaining_quantity}, Общая стоимость: {total_value} рублей")

        return all_products
